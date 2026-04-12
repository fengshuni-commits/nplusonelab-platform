#!/usr/bin/env python3
"""
N+1 LAB BOM 数据导入脚本
读取 Excel 文件并导入到 SQLite 数据库
"""

import sqlite3
import openpyxl
from pathlib import Path

# 配置
DB_PATH = '/root/.openclaw/workspace/projects/n1-lab-automation/backend/n1_lab_v3.db'
BOM_FOLDER = Path('/root/.openclaw/workspace/xiaohongshu/products/BOM-0330')  # 更新为 BOM-0330

# 产品 SKU 映射（文件名 -> 数据库 SKU -> Sheet 名）
# 格式：{文件名关键词：{SKU: [sheet 名列表]}}
# 注意：有些 SKU 需要多个 sheet 合并（如 AP2402-SKU2 需要合并两个 sheet）
PRODUCT_SHEET_MAPPING = {
    'A2504 餐边柜': {
        'A2504-SKU1': ['餐边柜']
    },
    'AP2401': {  # 置物架
        'AP2401-SKU1': ['单个框架'],
        'AP2401-SKU2': ['海洋板置物单元'],
        'AP2401-SKU3': ['海洋板抽屉'],
        'AP2401-SKU4': ['不锈钢展示板'],
        'AP2401-SKU5': ['不锈钢层板 '],  # 注意 sheet 名有空格
        'AP2401-SKU6': ['玻璃层板'],
        'AP2401-SKU7': ['织物挂杆']
    },
    'AP2402': {  # 推柜
        'AP2402-SKU1': ['展示推柜（不含侧板）'],
        'AP2402-SKU2': ['展示推柜（不含侧板）', '长虹亚克力侧板']  # 需要合并两个 sheet
    },
    'AP2403': {  # 茶几
        'AP2403-SKU1': ['桌架'],
        'AP2403-SKU2': ['桌架', '玻璃']  # 需要合并两个 sheet
    },
    'AP2501': {  # 双层推车
        'AP2501-SKU1': ['推车'],
        'AP2501-SKU2': ['推车']  # SKU2 和 SKU1 共用同一个 sheet（需要后续手动补充）
    },
    'AP2502': {  # 天地撑书架
        'AP2502-SKU1': ['天地杆'],
        'AP2502-SKU2': ['C型搁板'],  # 独立 SKU
        'AP2502-SKU3': ['天地杆+C型搁板'],  # 组合 SKU（注意 sheet 名是"天地杆+2C型搁板"）
        'AP2502-SKU4': ['一字板'],
        'AP2502-SKU5': ['工字盒']
    },
    'AP2503': {  # 铝型材桌
        'AP2503-SKU1': ['1800书桌'],  # 无空格！
        'AP2503-SKU2': ['1800书桌', '桌面']  # 合并
    }
}

def connect_db():
    """连接数据库"""
    conn = sqlite3.connect(DB_PATH)
    return conn

def get_product_id(cursor, sku_code):
    """根据 SKU 获取产品 ID"""
    cursor.execute("SELECT id FROM products WHERE sku_code = ?", (sku_code,))
    result = cursor.fetchone()
    return result[0] if result else None

def read_single_sheet(ws, sheet_name):
    """读取单个 sheet 的 BOM 数据"""
    bom_items = []
    
    # 提取 BOM 数据（从第 2 行开始，第 1 行是表头）
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # 跳过表头（第 1 行）
        if row_idx == 1:
            continue
        
        # 转换为字符串并清理
        cells = [str(cell) if cell is not None else '' for cell in row]
        
        # 跳过全空行
        if not any(c.strip() for c in cells):
            continue
        
        # 标准结构：[空，名称，图片，参数，数目，单位，采购数量，链接，参考单价，总价，备注]
        # 跳过表头行（第 1 行）
        if row_idx == 1:
            continue
        
        # 提取数据（固定列索引）
        part_name = cells[1].strip() if len(cells) > 1 else ''
        
        # 跳过空行或非零件行
        if not part_name or part_name in ['', '运费', '价格小计', '备注']:
            continue
        
        specs = cells[3].strip() if len(cells) > 3 else ''
        
        # 读取"数目"列（列 4）作为零件数量
        try:
            part_quantity_val = float(cells[4]) if len(cells) > 4 and cells[4] not in ['', 'None', 'nan'] else 1.0
            part_quantity = str(int(part_quantity_val)) if part_quantity_val == int(part_quantity_val) else str(part_quantity_val)
        except:
            part_quantity = '1'
        
        # 读取"采购数量"列（列 6）
        try:
            purchase_quantity_val = float(cells[6]) if len(cells) > 6 and cells[6] not in ['', 'None', 'nan'] else part_quantity_val
            purchase_quantity = str(int(purchase_quantity_val)) if purchase_quantity_val == int(purchase_quantity_val) else str(purchase_quantity_val)
        except:
            purchase_quantity = part_quantity
        
        link = cells[7].strip() if len(cells) > 7 else ''
        cost_str = cells[8].strip() if len(cells) > 8 else '0'  # 参考单价（列 8）
        total_cost_str = cells[9].strip() if len(cells) > 9 else '0'  # 总价列（列 9，含运费）
        remark = cells[10].strip() if len(cells) > 10 else ''
        
        # 转换成本为数字
        try:
            cost = float(cost_str) if cost_str else 0.0
        except:
            cost = 0.0
        
        # 转换总价为数字（含运费）
        try:
            total_cost = float(total_cost_str) if total_cost_str else (cost * purchase_quantity_val)
        except:
            total_cost = cost * purchase_quantity_val
        
        # 导入所有采购数量有值的行（不管备注是什么）
        if purchase_quantity_val is not None:
            bom_items.append({
                'part_name': part_name,
                'specs': specs,
                'quantity': part_quantity,
                'purchase_quantity': purchase_quantity,
                'link': link,
                'estimated_cost': cost,
                'total_cost': total_cost,  # 总价（含运费）
                'remark': remark,
                'is_kit': 'N+1' in remark and '零件包' in remark if remark else False
            })
    
    return bom_items

def read_bom_excel(file_path):
    """读取 Excel 文件中的 BOM 数据（旧函数，保留兼容）"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        print(f"  📄 文件：{file_path.name}")
        print(f"     Sheet 列表：{wb.sheetnames}")
        
        # 跳过第一个汇总 sheet，读取第二个明细 sheet
        if len(wb.sheetnames) < 2:
            print(f"     ⚠️ 只有一个 sheet，跳过")
            return []
        
        # 读取第二个 sheet（明细表）
        ws = wb[wb.sheetnames[1]]
        print(f"     使用 Sheet: {wb.sheetnames[1]}")
        
        # 提取 BOM 数据（从第 2 行开始，第 1 行是表头）
        bom_items = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # 跳过表头（第 1 行）
            if row_idx == 1:
                continue
            
            # 转换为字符串并清理
            cells = [str(cell) if cell is not None else '' for cell in row]
            
            # 跳过全空行
            if not any(c.strip() for c in cells):
                continue
            
            # 提取数据（根据 BOM-0330 的新结构）
            # 列 0: 空，列 1: 名称，列 2: 图片，列 3: 参数，列 4: 数目，列 5: 单位，列 6: 采购数量，列 7: 链接，列 8: 参考单价，列 9: 总价，列 10: 备注，列 11: 核价
            part_name = cells[1].strip() if len(cells) > 1 else ''
            
            # 跳过空行或非零件行（如运费、价格小计）
            if not part_name or part_name in ['', '运费', '价格小计', '备注']:
                continue
            
            specs = cells[3].strip() if len(cells) > 3 else ''
            # 读取"数目"列（列 4）作为零件数量
            try:
                part_quantity_val = float(cells[4]) if len(cells) > 4 and cells[4] not in ['', 'None', 'nan'] else 1.0
                part_quantity = str(int(part_quantity_val)) if part_quantity_val == int(part_quantity_val) else str(part_quantity_val)
            except:
                part_quantity = '1'
            
            # 读取"采购数量"列（列 6）作为采购用量
            try:
                purchase_quantity_val = float(cells[6]) if len(cells) > 6 and cells[6] not in ['', 'None', 'nan'] else part_quantity_val
                purchase_quantity = str(int(purchase_quantity_val)) if purchase_quantity_val == int(purchase_quantity_val) else str(purchase_quantity_val)
            except:
                purchase_quantity = part_quantity
            link = cells[7].strip() if len(cells) > 7 else ''
            cost_str = cells[8].strip() if len(cells) > 8 else '0'
            remark = cells[10].strip() if len(cells) > 10 else ''
            
            # 转换成本为数字
            try:
                cost = float(cost_str) if cost_str else 0.0
            except:
                cost = 0.0
            
            # 导入所有有淘宝/天猫链接的零件
            if link and ('taobao' in link or 'tmall' in link or link.startswith('http')):
                # 判断是否是零件包内的零件
                is_kit = 'N+1' in remark and '零件包' in remark
                # 打印调试信息
                if not is_kit:
                    print(f"     发现非零件包零件：{part_name}, 备注='{remark}', 链接={link[:50]}...")
                # 保存原始备注（不要覆盖）
                bom_items.append({
                    'part_name': part_name,
                    'specs': specs,
                    'quantity': part_quantity,
                    'purchase_quantity': purchase_quantity,
                    'link': link,
                    'estimated_cost': cost,
                    'remark': remark if remark else ('N+1 零件包' if is_kit else ''),  # 保留原始备注
                    'is_kit': is_kit
                })
        
        print(f"     ✅ 提取到 {len(bom_items)} 条 BOM 数据")
        return bom_items
        
    except Exception as e:
        print(f"  ❌ 读取失败：{e}")
        import traceback
        traceback.print_exc()
        return []

def import_bom_data():
    """主导入函数"""
    print("=" * 60)
    print("🔄 N+1 LAB BOM 数据导入")
    print("=" * 60)
    
    conn = connect_db()
    cursor = conn.cursor()
    
    # 1. 清理旧数据
    print("\n1️⃣ 清理旧 BOM 数据...")
    cursor.execute("SELECT COUNT(*) FROM bom_items")
    old_count = cursor.fetchone()[0]
    print(f"   删除前：{old_count} 条")
    
    cursor.execute("DELETE FROM bom_items")
    conn.commit()
    print(f"   ✅ 已清理 {old_count} 条旧数据")
    
    # 2. 导入新数据 - 遍历所有 SKU
    print("\n2️⃣ 导入新 BOM 数据（多 SKU 支持）...")
    total_imported = 0
    
    for excel_file in BOM_FOLDER.glob("*.xlsx"):
        print(f"\n📂 处理：{excel_file.name}")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            # 查找匹配的文件关键词
            file_key = None
            for key in PRODUCT_SHEET_MAPPING.keys():
                if key in excel_file.name:
                    file_key = key
                    break
            
            if not file_key:
                print(f"   ⚠️ 未找到 SKU 映射，跳过")
                continue
            
            sku_sheets = PRODUCT_SHEET_MAPPING[file_key]
            print(f"   找到 {len(sku_sheets)} 个 SKU 配置")
            print(f"   实际 Sheet 列表：{wb.sheetnames}")
            
            # 遍历每个 SKU
            for sku_code, sheet_names in sku_sheets.items():
                # 获取产品 ID
                product_id = get_product_id(cursor, sku_code)
                if not product_id:
                    print(f"   ⚠️ 产品 {sku_code} 不存在，跳过")
                    continue
                
                print(f"\n   📦 {sku_code} (ID:{product_id})")
                
                # 读取该 SKU 对应的所有 sheet
                all_bom_items = []
                for sheet_name in sheet_names:
                    # 尝试精确匹配或模糊匹配
                    actual_sheet = None
                    if sheet_name in wb.sheetnames:
                        actual_sheet = sheet_name
                    else:
                        # 尝试模糊匹配（包含关键词）
                        for ws_name in wb.sheetnames:
                            if sheet_name in ws_name or ws_name in sheet_name:
                                actual_sheet = ws_name
                                print(f"     ℹ️ 模糊匹配：'{sheet_name}' → '{actual_sheet}'")
                                break
                    
                    if not actual_sheet:
                        print(f"     ⚠️ Sheet '{sheet_name}' 不存在，跳过")
                        continue
                    
                    ws = wb[actual_sheet]
                    bom_items = read_single_sheet(ws, actual_sheet)
                    if bom_items:
                        all_bom_items.extend(bom_items)
                        print(f"     ✅ {sheet_name}: {len(bom_items)} 条")
                
                # 插入数据库
                if all_bom_items:
                    for item in all_bom_items:
                        cursor.execute("""
                            INSERT INTO bom_items (product_id, part_name, specs, quantity, purchase_quantity, link, estimated_cost, total_cost, remark)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            product_id,
                            item['part_name'],
                            item['specs'],
                            item['quantity'],
                            item['purchase_quantity'],
                            item['link'],
                            item['estimated_cost'],
                            item.get('total_cost', 0),  # 总价（含运费）
                            item.get('remark', '')
                        ))
                        total_imported += 1
                    print(f"     ✅ 共导入 {len(all_bom_items)} 条 BOM")
                else:
                    print(f"     ⚠️ 无 BOM 数据")
        
        except Exception as e:
            print(f"   ❌ 处理失败：{e}")
            import traceback
            traceback.print_exc()
    
    conn.commit()
    
    # 3. 验证结果
    print("\n3️⃣ 验证导入结果...")
    cursor.execute("SELECT COUNT(*) FROM bom_items")
    new_count = cursor.fetchone()[0]
    print(f"   导入后：{new_count} 条")
    print(f"   ✅ 成功导入 {total_imported} 条新数据")
    
    # 按产品统计
    print("\n📦 按产品统计:")
    cursor.execute("""
        SELECT p.sku_code, p.product_name, COUNT(b.id) as bom_count
        FROM products p
        LEFT JOIN bom_items b ON p.id = b.product_id
        GROUP BY p.id
        ORDER BY p.sku_code
    """)
    for row in cursor.fetchall():
        print(f"   {row[0]} ({row[1]}): {row[2]} 条 BOM")
    
    # 查看样例
    print("\n🔍 数据样例（前 5 条）:")
    cursor.execute("""
        SELECT p.sku_code, b.part_name, b.specs, b.quantity, b.link
        FROM bom_items b
        JOIN products p ON b.product_id = p.id
        LIMIT 5
    """)
    for row in cursor.fetchall():
        link_preview = row[4][:50] if row[4] else ''
        print(f"   {row[0]} | {row[1]} | 规格：{row[2]} | 数量：{row[3]} | 链接：{link_preview}...")
    
    conn.close()
    print("\n" + "=" * 60)
    print("✅ BOM 数据导入完成！")
    print("=" * 60)

if __name__ == '__main__':
    import_bom_data()
